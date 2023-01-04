# Install all the following required libraries in command prompt or terminal
# Standard Imports

from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
import os
from dotenv import load_dotenv


# Make sure to create a `.env` file to run the following line of code
load_dotenv()

workbook = openpyxl.load_workbook("participants.xlsx")
sheet = workbook.active

email_sender = "ai_ds.aicolegion@ves.ac.in"
email_password = os.getenv("mail_passcode")
em = EmailMessage()
subj = "AI CoLegion's Hack-AI-Thon: Participation Certificate"

def html_text():
    em.add_alternative(
        f"""\
        <!DOCTYPE html>
        <html>
            <body>
                    <p>Hello {name}!</p>
                    <br>
                    Thank you for participating in this hackathon organised by <b>AI CoLegion</b> in association with <b>GeeksforGeeks</b>.
                    <br>
                    With over 80+ teams, the event was a great success and we are grateful for your presence and enthusiasm. Here's the certificate of participation, you will find it attached with the mail.
                    <br>
                    <br>
                    ---
                    <br>
                    Regards,
                    <br>
                    Team AI CoLegion
                    <br>
                    <br>
                    <a href="https://www.linkedin.com/company/ai-colegion-vesit/">
                                    <img src="https://upload.wikimedia.org/wikipedia/commons/thumb/8/81/LinkedIn_icon.svg/120px-LinkedIn_icon.svg.png" alt="LinkedIn" style="width:40px; height:40px;" />
                    &nbsp;
                    </a>
                    <a href="https://www.instagram.com/aicolegion_vesit/"> <img src="https://i.imgur.com/l6kR6Av.png" alt ="Instagram" style="width: 40px; height:40px;" /> </a>
                </p>
            </body>
        </html>
    """,
    subtype="html",
)

em["From"] = email_sender
em["Subject"] = subj
context = ssl.create_default_context()
# email_receiver = '2020.naresh.shewkani@ves.ac.in'
for i in range(2,215):    #first index is inclusive, last index is exclusive
    if sheet.cell(row=i, column=2).value is None:
        pass
    else:
        email_receiver = sheet.cell(row=i,column=2).value
        name = sheet.cell(row=i,column=1).value
        html_text()
        pdf_path = f'generated_pdf/{name}.pdf' 
        with open(pdf_path, 'rb') as f:
            file_data = f.read()
            file_name = name + '.pdf'
        em.add_attachment(file_data, maintype='application',subtype='octet-stream',filename=file_name)
        em['To'] = email_receiver
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            # with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            #     smtp.starttls()
            smtp.login(email_sender, email_password)
            smtp.sendmail(email_sender, email_receiver, em.as_string())
            print(f'Message sent to {name} with the index {i}')
            em.clear_content()
        del em['To']